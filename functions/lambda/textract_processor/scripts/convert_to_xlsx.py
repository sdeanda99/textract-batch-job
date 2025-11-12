#!/usr/bin/env python3
"""
Convert Textract JSON results to XLSX format for Google Sheets/Excel import
Uses openpyxl for better handling of complex data
"""

import json
import os
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl not installed!")
    print("Install with: pip install openpyxl")
    print("Then run: python3 ./scripts/convert_to_xlsx.py")
    exit(1)

def process_json_file(json_path):
    """Extract key information from a Textract JSON result"""
    with open(json_path, 'r') as f:
        data = json.load(f)

    # Extract metadata
    metadata = data.get('metadata', {})

    # Extract all text
    raw_text_blocks = data.get('raw_text', [])
    full_text = '\n'.join([block.get('text', '') for block in raw_text_blocks])

    # Extract key-value pairs
    key_value_pairs = data.get('key_value_pairs', [])

    # Extract tables
    tables = data.get('tables', [])

    return {
        'source_file': metadata.get('source_file', ''),
        'batch': metadata.get('batch', ''),
        'job_id': metadata.get('job_id', ''),
        'processed_time': metadata.get('processed_time', ''),
        'total_blocks': metadata.get('total_blocks', 0),
        'full_text': full_text,
        'key_value_pairs': key_value_pairs,
        'tables': tables,
        'stats': {
            'text_blocks': len(raw_text_blocks),
            'kv_pairs': len(key_value_pairs),
            'tables': len(tables)
        }
    }

def create_summary_sheet(wb, results):
    """Create summary sheet with high-level info"""
    ws = wb.active
    ws.title = "Summary"

    # Headers
    headers = ['Source File', 'Batch', 'Processed Time', 'Text Blocks', 'Key-Value Pairs', 'Tables', 'Text Preview']
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    # Data rows
    for result in results:
        text_preview = result['full_text'][:200] + '...' if len(result['full_text']) > 200 else result['full_text']
        ws.append([
            result['source_file'],
            result['batch'],
            result['processed_time'],
            result['stats']['text_blocks'],
            result['stats']['kv_pairs'],
            result['stats']['tables'],
            text_preview
        ])

    # Auto-size columns
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = max(
            len(str(ws.cell(row, col).value or ''))
            for row in range(1, ws.max_row + 1)
        )
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

def create_key_values_sheet(wb, results):
    """Create sheet with all key-value pairs"""
    ws = wb.create_sheet("Key-Value Pairs")

    # Headers
    headers = ['Source File', 'Batch', 'Key', 'Value', 'Confidence']
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for result in results:
        for kv in result['key_value_pairs']:
            ws.append([
                result['source_file'],
                result['batch'],
                kv.get('key', ''),
                kv.get('value', ''),
                round(kv.get('confidence', 0), 2)
            ])

    # Auto-size columns
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 30

def create_full_text_sheet(wb, results):
    """Create sheet with full extracted text"""
    ws = wb.create_sheet("Full Text")

    # Headers
    headers = ['Source File', 'Batch', 'Full Text']
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    header_font = Font(bold=True, color="000000")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for result in results:
        ws.append([
            result['source_file'],
            result['batch'],
            result['full_text']
        ])

    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 100

def create_tables_sheet(wb, results):
    """Create sheet with extracted tables"""
    ws = wb.create_sheet("Tables")

    # Headers
    headers = ['Source File', 'Batch', 'Table Number', 'Table Data (JSON)', 'Confidence']
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="E74856", end_color="E74856", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(1, col)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for result in results:
        for idx, table in enumerate(result['tables'], 1):
            # Convert table rows to readable format
            table_text = '\n'.join([' | '.join(row) for row in table.get('rows', [])])

            ws.append([
                result['source_file'],
                result['batch'],
                idx,
                table_text,
                round(table.get('confidence', 0), 2)
            ])

    # Set column widths
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 80
    ws.column_dimensions['E'].width = 12

def main():
    results_dir = Path('../textract_results')
    output_xlsx = '../textract_results.xlsx'

    if not results_dir.exists():
        print("Error: textract_results directory not found!")
        print("Run ./download_results.sh first")
        return

    # Find all JSON files
    json_files = list(results_dir.rglob('*.json'))

    if not json_files:
        print("No JSON files found in textract_results/")
        return

    print(f"Processing {len(json_files)} JSON files...")

    # Process all files
    results = []
    for idx, json_file in enumerate(json_files, 1):
        try:
            result = process_json_file(json_file)
            results.append(result)
            if idx % 50 == 0:
                print(f"  Processed {idx}/{len(json_files)} files...")
        except Exception as e:
            print(f"Error processing {json_file}: {e}")

    if not results:
        print("No results to write!")
        return

    print(f"\nCreating Excel workbook with {len(results)} documents...")

    # Create workbook
    wb = Workbook()

    # Create sheets
    print("  Creating Summary sheet...")
    create_summary_sheet(wb, results)

    print("  Creating Key-Value Pairs sheet...")
    create_key_values_sheet(wb, results)

    print("  Creating Full Text sheet...")
    create_full_text_sheet(wb, results)

    print("  Creating Tables sheet...")
    create_tables_sheet(wb, results)

    # Save workbook
    wb.save(output_xlsx)

    print(f"\n✓ Excel file created: {output_xlsx}")
    print(f"\nWorkbook contains:")
    print(f"  - Summary: High-level overview of all {len(results)} documents")
    print(f"  - Key-Value Pairs: All extracted form fields")
    print(f"  - Full Text: Complete text extraction")
    print(f"  - Tables: All extracted tables")
    print(f"\nYou can now:")
    print(f"  1. Open in Excel/Numbers")
    print(f"  2. Import to Google Sheets")
    print(f"  3. Use with n8n Excel/Google Sheets nodes")

if __name__ == '__main__':
    main()
